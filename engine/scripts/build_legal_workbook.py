"""Render the legal-review workbook (xlsx) from the canonical payload — reproducibly.

All sheet data comes from scripts/export_legal_review_payload.build_payload()
(NEW/KNOWN/Absence findings, recall misses, zone-3 scores, criteria, master
rows, refuter verdicts when a full-indicator-v2 run exists). This script only
renders + styles, and optionally assembles the send-ready package.

Usage:
  .venv/bin/python scripts/build_legal_workbook.py            # workbook only
  .venv/bin/python scripts/build_legal_workbook.py --package  # + outputs/legal_send/
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

RUNS = ["final_si_p6", "final_si_p7", "final_ma_p6", "final_ma_p7",
        "final_au_p6", "final_au_p7"]
OUT = Path("outputs/legal_recall_review/ClauseChain_Legal_Review_Workbook.xlsx")

HEAD_FONT = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="0FB5A7")
WRAP = Alignment(wrap_text=True, vertical="top")

REVIEW_COLS = ["Reviewer decision", "Reviewer correction/reasoning",
               "Reviewer official source URL", "Reviewer name", "Reviewer role",
               "Review date", "Citation checked", "Mapping checked", "Currentness checked"]


def _sheet(wb, title, headers, widths):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"
    return ws


def _indicator_questions() -> dict[str, dict]:
    out = {}
    for p in ("6", "7"):
        cfg = yaml.safe_load(Path(f"configs/rdtii/pillar_{p}.yaml").read_text())
        for ind_id, ind in (cfg.get("indicators") or {}).items():
            out[ind_id] = ind if isinstance(ind, dict) else {}
    return out


def _alignment_label(finding: dict) -> str:
    proof = finding.get("citation_proof") or {}
    method = str(proof.get("alignment_status") or proof.get("method")
                 or proof.get("alignment") or "").lower()
    score = proof.get("alignment_score")
    if "anchor" in method or (finding.get("Location Reference") or "").startswith("#"):
        return f"anchor ({score if score is not None else 1})"
    if "exact" in method:
        return f"exact ({score if score is not None else 1})"
    return "unaligned (0)"


def _display(value, limit: int) -> str:
    if value is None:
        return ""
    text = value if isinstance(value, str) else json.dumps(value, ensure_ascii=False)
    return text[:limit]


def _surrounding_context(finding: dict) -> str:
    return _display(finding.get("raw_context") or finding.get("Notes") or "", 1400)


def _official_urls(entry: dict) -> str:
    values = entry.get("references") or entry.get("urls") or entry.get("url") or []
    if isinstance(values, str):
        values = [values]
    return "\n".join(str(value).strip() for value in values if str(value).strip())[:900]


def _recall_rationale(miss: dict) -> str:
    evidence = miss.get("evidence") or {}
    technical = evidence.get("technical_class") or miss.get("class") or "UNCLASSIFIED"
    emitted = ", ".join(miss.get("emitted_under") or [])
    messages = {
        "NOT_IN_CORPUS": (
            "The master citation was not located in the rebuilt eligible corpus. Review whether "
            "the official instrument/version was acquired and whether the master reference is current."
        ),
        "IN_CORPUS_NOT_EMITTED": (
            "The cited provision exists in the corpus, but no surviving finding was emitted for the "
            "master indicator. Review legal mapping, context and deterministic gate outcomes."
        ),
        "EMITTED_OTHER_INDICATOR": (
            f"The cited provision was emitted under {emitted or 'another indicator'}, not under the "
            "master indicator. Decide whether the master mapping, engine mapping, or both need correction."
        ),
        "GOLD_REF_UNPARSEABLE": (
            "The master citation could not be parsed into a stable legal anchor. Verify the citation "
            "against the official instrument and provide a corrected reference if necessary."
        ),
    }
    detail = messages.get(technical, f"Technical classification: {technical}. Legal review is required.")
    return f"{detail} Proposed system verdict: {miss.get('proposed_verdict') or 'REVIEW_REQUIRED'}."


def _col_width(header: str) -> int:
    h = header.lower()
    if any(k in h for k in ("snippet", "context", "rationale", "manifest", "reasoning",
                            "impact", "guidance", "question", "test", "criteria")):
        return 46
    if any(k in h for k in ("law", "instrument", "url", "act", "evidence", "warning", "reason")):
        return 28
    if h.endswith(" id") or h in ("indicator", "economy"):
        return 10
    return 15


def render_workbook(payload: dict, out: Path = OUT) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    counts = payload.get("counts", {})
    ws = _sheet(wb, "Instructions", ["ClauseChain Legal Review", ""], [34, 116])
    ws.append(["Generated by", "scripts/build_legal_workbook.py (payload: export_legal_review_payload.py) — "
               "regenerate any time; never hand-edit structure"])
    ws.append(["Review order", f"1) NEW Findings ({counts.get('new')}) · 2) Absence Review ({counts.get('absence')}) · "
               f"3) Recall Misses ({counts.get('recall')}) · 4) Zone-3 Scores ({counts.get('zone3')}) · "
               f"5) KNOWN Findings ({counts.get('known')})"])
    ws.append(["Refuter", payload.get("refuter_status", "")])
    ws.append(["Finding decisions", "approve | reject | needs-correction — reviewer name, role and date are mandatory; "
               "a row without an explicit approve is excluded from the final export."])
    ws.append(["Recall verdicts", "REAL_MISS | GOLD_WRONG | GOLD_AMBIGUOUS | CORRECT_ABSTENTION | NEEDS_CORRECTION. "
               "Malaysia rows: ESCAP planted deliberate errors (confirmed) — GOLD_WRONG with evidence earns points; "
               "cross-check the 124-finding error audit."])
    ws.append(["Method", "Use Indicator Criteria as the controlling methodology; open each official source URL and "
               "compare the exact quotation via the Source Proof Bundle (review/index.html)."])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP

    for title, sheet in payload.get("sheets", {}).items():
        headers, rows = sheet.get("headers", []), sheet.get("rows", [])
        ws = _sheet(wb, title, headers, [_col_width(h) for h in headers])
        for row in rows:
            ws.append(["" if v is None else (v if isinstance(v, (int, float)) else str(v))
                       for v in row])
        for r in ws.iter_rows(min_row=2):
            for c in r:
                c.alignment = WRAP

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def build_package(payload: dict, workbook: Path) -> Path:
    """Assemble outputs/legal_send/: workbook + proof-bundle zip + manifest/readme."""
    import hashlib
    import shutil
    import zipfile as _zip

    send = Path("outputs/legal_send")
    send.mkdir(parents=True, exist_ok=True)
    wb_dst = send / workbook.name
    shutil.copy2(workbook, wb_dst)

    bundle = send / "ClauseChain_Source_Proof_Bundle.zip"
    with _zip.ZipFile(bundle, "w", _zip.ZIP_DEFLATED) as z:
        for path in sorted(Path("submission/review").rglob("*")):
            if path.is_file() and path.name != ".DS_Store":
                z.write(path, path.relative_to("submission"))

    def sha(p: Path) -> str:
        return hashlib.sha256(p.read_bytes()).hexdigest()

    counts = payload.get("counts", {})
    (send / "00_READ_ME_FIRST.txt").write_text(f"""CLAUSECHAIN — LEGAL REVIEW PACKAGE

Please open:
  1. {wb_dst.name}
  2. {bundle.name} (unzip it, then open review/index.html)

REVIEW ORDER
  1. NEW Findings — {counts.get('new')} proposed discoveries (refuter verdicts are advisory; you decide)
  2. Absence Review — {counts.get('absence')} no-evidence/score-zero candidates
  3. Recall Misses — {counts.get('recall')} ESCAP master anchors requiring adjudication
  4. Zone-3 Scores — {counts.get('zone3')} indicator scores
  5. KNOWN Findings — {counts.get('known')} surviving baseline evidence rows

HOW TO REVIEW
  - Use the Indicator Criteria sheet as the controlling methodology.
  - Open the official source URL and compare the exact quotation and context.
  - Finding rows: approve, reject, or needs-correction.
  - Recall rows: REAL_MISS, GOLD_WRONG, GOLD_AMBIGUOUS, CORRECT_ABSTENTION, NEEDS_CORRECTION.
  - Complete reviewer name, role and review date on every decided row.
  - For an approved finding set citation, mapping and currentness checks to "yes".

REFUTER NOTE
  {payload.get('refuter_status', '')}

Nothing in this package is pre-approved. Reviewer decisions remain blank.
""", encoding="utf-8")

    (send / "01_PACKAGE_MANIFEST.txt").write_text(f"""CLAUSECHAIN LEGAL REVIEW PACKAGE — MANIFEST (generated)

Evidence rows: NEW {counts.get('new')} | KNOWN {counts.get('known')} | Absence {counts.get('absence')}
Decisions: Recall {counts.get('recall')} | Zone-3 {counts.get('zone3')} | Master reference rows {counts.get('master')}

SHA-256
  {sha(wb_dst)}  {wb_dst.name}
  {sha(bundle)}  {bundle.name}

Regenerate: .venv/bin/python scripts/build_legal_workbook.py --package
Verify engine: .venv/bin/python -m pytest tests/ -q
This is a review package, not an approved submission.
""", encoding="utf-8")
    return send


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--package", action="store_true",
                        help="also assemble outputs/legal_send/ (zip + manifest)")
    args = parser.parse_args()
    # lazy import — the exporter imports helpers from this module
    from scripts.export_legal_review_payload import build_payload

    payload = build_payload()
    out = render_workbook(payload)
    counts = payload.get("counts", {})
    print(f"wrote {out} — " + ", ".join(f"{k}:{v}" for k, v in sorted(counts.items())))
    if args.package:
        send = build_package(payload, out)
        print(f"package assembled: {send}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
